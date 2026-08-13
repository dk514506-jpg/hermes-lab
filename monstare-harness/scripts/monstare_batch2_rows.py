#!/usr/bin/env python3
"""Print batch-2 rows from the workbook directly (full columns) + audit verdicts."""
import csv
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
TARGET = ["A9-01", "A9-02", "A9-04", "A9-05", "A9-06", "CORE-09", "CORE-10", "CORE-15"]

wb = load_workbook(XLSX, data_only=True)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows(values_only=True))
header = [str(c).strip() if c is not None else "" for c in rows[0]]
ci = {h: i for i, h in enumerate(header)}

verdicts = {}
with open("/opt/data/Monstare_source_link_audit_2026-08-13.csv") as f:
    for r in csv.DictReader(f):
        verdicts[(r["RowID"], r["URL_Kind"])] = (r["Verdict"], r["HTTP_Code"], r["Notes"][:70])

for r in rows[1:]:
    rid = str(r[ci["ID"]]).strip() if r[ci["ID"]] else ""
    if rid not in TARGET:
        continue
    print("=" * 100)
    print(f"{rid} — {r[ci['Citation']]}")
    for k in ["Access Type", "Source Discovery Status", "Pri", "Area", "Function", "Evidence Type",
              "Domain", "Causal Status", "Cosmo Rel.", "H1", "H2", "Fail Mode", "Artifact Affected", "Open Charge"]:
        print(f"  {k}: {r[ci[k]]}")
    print(f"  Readable: {r[ci['Readable Source URL']]}")
    print(f"    audit: {verdicts.get((rid,'readable'))}")
    print(f"  Landing: {r[ci['Source Landing URL']]}")
    print(f"    audit: {verdicts.get((rid,'landing'))}")
    print(f"  Discovery Notes: {str(r[ci['Discovery Notes']])[:200]}")
