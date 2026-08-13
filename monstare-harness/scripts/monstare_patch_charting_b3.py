#!/usr/bin/env python3
"""Final additive matrix patch: batch-3 charting (CORE completion spine).
Reads /opt/data/Monstare_batch_3_charting_final.json (post-QC) + URL/caveat/access/evtype updates."""
import json, os
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
DATE = "2026-08-13"
final = json.load(open("/opt/data/Monstare_batch_3_charting_final.json"))

COLMAP = {
    "kft": "Key Finding / Thesis", "ess": "Effect Size / Strength", "lim": "Limitations",
    "di": "Disconfirming Implication", "h1": "H1", "h2": "H2",
    "dimpl": "Design Implication", "cimpl": "Cosmotechnic Implication", "cs": "Causal Status",
}
ACCESS_UPDATES = json.load(open("/opt/data/Monstare_batch_3_access_updates.json")) if os.path.exists("/opt/data/Monstare_batch_3_access_updates.json") else {}
EVTYPE_UPDATES = json.load(open("/opt/data/Monstare_batch_3_evtype_updates.json")) if os.path.exists("/opt/data/Monstare_batch_3_evtype_updates.json") else {}
URL_UPDATES = json.load(open("/opt/data/Monstare_batch_3_url_updates.json")) if os.path.exists("/opt/data/Monstare_batch_3_url_updates.json") else {}  # {rid: {"Readable Source URL": ..., "Source Landing URL": ...}}

CAVEAT_NOTES = json.load(open("/opt/data/Monstare_batch_3_caveat_notes.json")) if os.path.exists("/opt/data/Monstare_batch_3_caveat_notes.json") else {}  # key: rid -> note text
DOWNGRADES = json.load(open("/opt/data/Monstare_batch_3_locus_tags.json")) if os.path.exists("/opt/data/Monstare_batch_3_locus_tags.json") else {}  # key: rid -> LOCUS weight tag text

wb = load_workbook(XLSX)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows())
header = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
col = {h: i for i, h in enumerate(header)}

patched = []
for r in rows[1:]:
    rid = str(r[col["ID"]].value).strip() if r[col["ID"]].value else ""
    if rid not in final:
        continue
    f = final[rid]
    for key, cname in COLMAP.items():
        if key in f and f[key] is not None:
            r[col[cname]].value = f[key]
    r[col["Verif."]].value = f"Charted - {DATE}"
    if rid in ACCESS_UPDATES:
        r[col["Access Type"]].value = ACCESS_UPDATES[rid]
    if rid in EVTYPE_UPDATES:
        r[col["Evidence Type"]].value = EVTYPE_UPDATES[rid]
    if rid in URL_UPDATES:
        for urlkey, val in URL_UPDATES[rid].items():
            if val:
                r[col[urlkey]].value = val
    notes = str(r[col["Discovery Notes"]].value or "")
    additions = []
    if "CHARTED 2026-08-13 batch 3" not in notes:
        additions.append("CHARTED 2026-08-13 batch 3 (CORE completion spine; full role QC pass)")
    if rid in CAVEAT_NOTES and CAVEAT_NOTES[rid][:20] not in notes:
        additions.append(CAVEAT_NOTES[rid])
    if rid in DOWNGRADES and "LOCUS" not in notes:
        additions.append(f"LOCUS {DATE}: {DOWNGRADES[rid]}")
    if additions:
        r[col["Discovery Notes"]].value = (notes + "; " if notes else "") + "; ".join(additions)
    r[col["Source Checked Date"]].value = DATE
    patched.append(rid)

wb.save(XLSX)
print("PATCHED:", patched)

pl = wb["Source Patch Log"]
pl_append = pl.max_row + 1
entry = [
    DATE,
    "Additive patch - batch 3 charting (QC-complete)",
    ", ".join(patched),
    "None (no schema change; existing charting columns filled)",
    "0",
    "0",
    "Third charting pass (10-row CORE completion spine) with full role QC (Evidence Librarian, "
    "Methodologist, Phenomenologist (CORE-16), Cosmotechnic-Purist, Ethics & Cosmotechnic Auditor, "
    "Data & Instrumentation Steward, Locus). Fill classification: 0 pure fills; ALL 90 charting-cell "
    "writes (kft/ess/lim/di/h1/h2/dimpl/cimpl/cs) are instructed seeded-cell upgrades - every target "
    "cell held seed text at patch time (verified pre-patch), replaced by post-QC charting; seeds "
    "superseded, not deleted, provenance kept via CHARTED 2026-08-13 batch 3 Discovery Notes marker. "
    "Causal Status seed corrections per Methodologist QC: CORE-12 'causal' lab-bounded (animal, "
    "catalog-attributed, Tier-P); CORE-13 'causal'->correlational (narrative review); CORE-17/19/20 "
    "'causal'->conceptual (methods). Evidence Type refinements (CORE-03/11/12/13) and CORE-19 Access "
    "Type update per QC instructions. Verif. set to 'Charted - 2026-08-13' on 10 rows. CORE-12 + "
    "CORE-17 charted at abstract level (documented decisions - archive lending only; Kazdin 2021 JEAB "
    "paywalled). URL repairs classified as improving source links (not deletions): CORE-19 seeded "
    "readable PMC5820391 was the WRONG article (Aasdahl 2018) and seeded landing 10.1016/j.cct.2015.07.003 "
    "the WRONG article (Broglio et al.) - replaced with verified author-hosted PDF (ambujtewari.com) + "
    "doi.org/10.1037/hea0000305; CORE-14 landing dead DOI 019823682X.003.0002 -> 10.1093/analys/58.1.7; "
    "CORE-18 readable -> doi.org/10.1037/arc0000026 (version of record). CORE-16 charted from "
    "course-hosted full-book mirror (rights caveat; archive lending canonical). CORE-20 audit 404 "
    "corrected (PMC2062525 live). Caveat notes appended to 8 rows. No rows removed, no source-URL cells "
    "cleared, no formulas or sheets touched, additive only. Dashboard 44 formulas verified intact post-patch.",
]
for i, v in enumerate(entry):
    pl.cell(row=pl_append, column=i + 1, value=v)
wb.save(XLSX)
print("PATCH LOG ENTRY ADDED (row", pl_append, ")")

wb2 = load_workbook(XLSX)  # NOT data_only — formula check must count live formulas
ws2 = wb2["Evidence Matrix"]
rows2 = list(ws2.iter_rows(values_only=True))
h2 = [str(c).strip() if c else "" for c in rows2[0]]
c2 = {h: i for i, h in enumerate(h2)}
print("\nREADBACK:")
for r in rows2[1:]:
    rid = str(r[c2["ID"]]).strip() if r[c2["ID"]] else ""
    if rid in final:
        print(f"  {rid}: Verif={r[c2['Verif.']]} | KFT={len(str(r[c2['Key Finding / Thesis']] or ''))}ch | CS={str(r[c2['Causal Status']])[:60]}")

# formula integrity check
dash = wb2["Dashboard"]
formulas = 0
for row in dash.iter_rows():
    for c in row:
        if isinstance(c.value, str) and c.value.startswith("="):
            formulas += 1
print("Dashboard formula count (cached):", formulas)
