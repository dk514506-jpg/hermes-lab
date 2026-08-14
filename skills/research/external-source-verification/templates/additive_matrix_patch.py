#!/usr/bin/env python3
"""Template: additive in-place patch of a canonical evidence matrix (openpyxl).

Copy and adapt to the target workbook. Principles (Silvey's law): the matrix is the
SINGLE canonical database — patch it in place, additively; never rebuild or fork it.

- Load WITHOUT data_only=True so formulas are preserved (openpyxl strips cached
  formula values on save; Excel recomputes on open — say so if no LibreOffice).
- Notes appends are idempotent: guarded on a marker so re-runs don't duplicate.
- One patch-log entry per pass.
- Read back the patched rows after saving.

Run: uv run --with openpyxl python3 additive_matrix_patch.py
"""
from openpyxl import load_workbook

XLSX = "path/to/canonical_matrix.xlsx"
SHEET = "Evidence Matrix"
LOG_SHEET = "Source Patch Log"
DATE = "2026-08-13"

# row_id -> (new_readable_url, new_landing_url_or_None, access_type, discovery_status, note)
PATCH = {
    # "CORE-06": (
    #     "https://example.org/replacement.pdf",
    #     "https://doi.org/10.xxxx/xxxxx",
    #     "Open PDF / DOI landing",
    #     "Located - readable source (replacement verified 2026-08-13)",
    #     "REPLACEMENT 2026-08-13: old URL dead (<reason>). Replacement text-layer verified.",
    # ),
}
MARKER = f"REPLACEMENT {DATE}"   # also handle "SOURCE ADDED {DATE}"

wb = load_workbook(XLSX)  # formulas preserved
ws = wb[SHEET]
rows = list(ws.iter_rows())
header = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
col = {h: i for i, h in enumerate(header)}

patched = []
for r in rows[1:]:
    rid = r[col["ID"]].value
    if rid is None or str(rid).strip() not in PATCH:
        continue
    rid = str(rid).strip()
    new_readable, new_landing, acc, st, note = PATCH[rid]
    notes_cell = r[col["Discovery Notes"]]
    old_notes = str(notes_cell.value).strip() if notes_cell.value else ""
    r[col["Readable Source URL"]].value = new_readable
    if new_landing:
        r[col["Source Landing URL"]].value = new_landing
    r[col["Access Type"]].value = acc
    r[col["Source Discovery Status"]].value = st
    if MARKER not in old_notes:                       # idempotent append
        notes_cell.value = (old_notes + "; " if old_notes else "") + note
    if "Source Checked Date" in col:
        r[col["Source Checked Date"]].value = DATE
    patched.append(rid)

print("PATCHED:", patched)

# One patch-log entry per pass (adjust columns to the log sheet's header)
pl = wb[LOG_SHEET]
entry = [DATE, "Additive patch - link replacements (audit-driven)",
         ", ".join(patched), "", "0", "0",
         "Replacements verified reachable + text-layer checked on " + DATE +
         ". No rows removed; no existing source cells removed; additive only."]
for i, v in enumerate(entry):
    pl.cell(row=pl.max_row + 1, column=i + 1, value=v)

wb.save(XLSX)
print("SAVED:", XLSX)

# Readback verification
wb2 = load_workbook(XLSX, data_only=True)
ws2 = wb2[SHEET]
h2 = [str(c).strip() if c else "" for c in next(ws2.iter_rows(min_row=1, max_row=1))]
c2 = {h: i for i, h in enumerate(h2)}
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[c2["ID"]] and str(r[c2["ID"]]).strip() in PATCH:
        print("  READBACK", str(r[c2["ID"]]).strip(), "->", r[c2["Readable Source URL"]])
