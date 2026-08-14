#!/usr/bin/env python3
"""Print batch-3 rows (CORE-03/11/12/13/14/16/17/18/19/20) from the workbook (full columns) + audit verdicts + Source Patch Log tail."""
import csv
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
TARGET = ["CORE-03", "CORE-11", "CORE-12", "CORE-13", "CORE-14", "CORE-16", "CORE-17", "CORE-18", "CORE-19", "CORE-20"]

wb = load_workbook(XLSX, data_only=True)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows(values_only=True))
header = [str(c).strip() if c is not None else "" for c in rows[0]]
ci = {h: i for i, h in enumerate(header)}
print("COLUMNS (%d):" % len(header))
print(" | ".join(header))
print()

verdicts = {}
with open("/opt/data/Monstare_source_link_audit_2026-08-13.csv") as f:
    for r in csv.DictReader(f):
        verdicts[(r["RowID"], r["URL_Kind"])] = (r["Verdict"], r["HTTP_Code"], r["Notes"][:60])

SHOW = ["Access Type", "Source Discovery Status", "Pri", "Area", "Function", "Evidence Type",
        "Domain", "Causal Status", "Cosmo Rel.", "H1", "H2", "Fail Mode", "Artifact Affected",
        "Open Charge", "Source Checked Date", "Verif.", "Staleness Status", "Re-search By"]
TEXT = ["Key Finding / Thesis", "Effect Size / Strength", "Limitations",
        "Disconfirming Implication", "Design Implication", "Cosmotechnic Implication"]

for r in rows[1:]:
    rid = str(r[ci["ID"]]).strip() if r[ci["ID"]] else ""
    if rid not in TARGET:
        continue
    print("=" * 110)
    print(f"{rid} — {r[ci['Citation']]}")
    for k in SHOW:
        print(f"  {k}: {r[ci[k]]}")
    print(f"  Readable: {r[ci['Readable Source URL']]}")
    print(f"    audit: {verdicts.get((rid,'readable'))}")
    print(f"  Landing: {r[ci['Source Landing URL']]}")
    print(f"    audit: {verdicts.get((rid,'landing'))}")
    print(f"  Discovery Notes: {str(r[ci['Discovery Notes']])[:400]}")
    print(f"  Suggested Search Query: {str(r[ci['Suggested Search Query']])[:150]}")
    print(f"  Scholar/Web Search URL: {str(r[ci['Scholar/Web Search URL']])[:150]}")
    for k in TEXT:
        v = r[ci[k]]
        if v is not None and str(v).strip():
            print(f"  [{k}] (SEEDED): {str(v)[:300]}")
        else:
            print(f"  [{k}]: <blank>")

print()
print("#" * 110)
print("SOURCE PATCH LOG TAIL")
ws2 = wb["Source Patch Log"]
for r in list(ws2.iter_rows(values_only=True))[-12:]:
    cells = [str(c)[:120] if c is not None else "" for c in r]
    print(" || ".join(cells))
