#!/usr/bin/env python3
"""Final additive matrix patch: batch-1 charting + access caveats + landing fix.
Reconciles Methodologist/Purist/Librarian/Ethics/Locus/Data-Steward outputs.
Verified: JSON carries cimpl for all 8 rows (Data Steward's null-claim was a misread)."""
import json
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
DATE = "2026-08-13"
final = json.load(open("/opt/data/Monstare_batch_1_charting_final.json"))

COLMAP = {
    "kft": "Key Finding / Thesis", "ess": "Effect Size / Strength", "lim": "Limitations",
    "di": "Disconfirming Implication", "h1": "H1", "h2": "H2",
    "dimpl": "Design Implication", "cimpl": "Cosmotechnic Implication", "cs": "Causal Status",
}

ACCESS_UPDATES = {
    "CORE-04": "Archive lending (1997 book, no open PDF) / companion 1994 chapter (open, albertbandura.com)",
    "CORE-05": "Open PDF but scanned - read via OCR text",
    "CORE-06": "Author-site PDF - broken direct text layer, readable only via OCR; DOI landing = 10.1111/1467-6419.00150 (publisher copy paywalled)",
    "CORE-08": "Open PDF = accepted manuscript (not typeset); cite Behavioural Public Policy 6(4) 2022",
    "A1-01": "Archive.org scan (Text PDF, rights caution) / publisher landing",
}
A1_LANDING = "https://www.urbanomic.com/book/question-concerning-technology-china/"
DOWNGRADES = {"CORE-01", "CORE-02", "CORE-04", "CORE-06"}  # Locus: supporting weight

wb = load_workbook(XLSX)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows())
header = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
col = {h: i for i, h in enumerate(header)}

patched = []
for r in rows[1:]:
    rid = str(r[col["ID"]].value).strip() if r[col["ID"]].value else ""
    if rid not in final:
        continue
    f = final[rid]
    for key, cname in COLMAP.items():
        if key in f and f[key] is not None:
            r[col[cname]].value = f[key]
    r[col["Verif."]].value = f"Charted - {DATE}"
    if rid in ACCESS_UPDATES:
        r[col["Access Type"]].value = ACCESS_UPDATES[rid]
    if rid == "A1-01":
        r[col["Source Landing URL"]].value = A1_LANDING
    notes = str(r[col["Discovery Notes"]].value or "")
    additions = []
    if "CHARTED 2026-08-13 batch 1" not in notes:
        additions.append("CHARTED 2026-08-13 batch 1 (motivational/cosmotechnic spine; full role QC pass)")
    if rid in DOWNGRADES and "LOCUS" not in notes:
        additions.append("LOCUS 2026-08-13: supporting weight (not load-bearing); empirical anchor delegated per charting")
    if rid == "A1-01" and "verified body claims" not in notes:
        additions.append("BODY CLAIMS VERIFIED 2026-08-13 from archive scan pp.25-50 (~ print pp.18-43): Heidegger-impasse, §1, §2 incl. cosmotechnics definition (scan p.37) and Leroi-Gourhan/Gille critiques; file A1-01_intro_verified.txt")
    if additions:
        r[col["Discovery Notes"]].value = (notes + "; " if notes else "") + "; ".join(additions)
    r[col["Source Checked Date"]].value = DATE
    patched.append(rid)

wb.save(XLSX)
print("PATCHED:", patched)

# ---- Source Patch Log (truthful entry per Data Steward audit, cimpl clause corrected) ----
pl = wb["Source Patch Log"]
pl_rows = list(pl.iter_rows(values_only=True))
pl_append = pl.max_row + 1
entry = [
    DATE,
    "Additive patch - batch 1 charting (QC-complete)",
    "CORE-01, CORE-02, CORE-04, CORE-05, CORE-06, CORE-08, A1-01, HUI-2024",
    "None (no schema change; existing charting columns filled)",
    "0",
    "0",
    "First charting pass with full role QC (Evidence Librarian, Methodologist, Cosmotechnic-Purist, "
    "Ethics & Cosmotechnic Auditor, Data & Instrumentation Steward, Locus). Citation-is-not-evidence "
    "respected; CORE-04 charted via 1994 companion chapter (1997 book is archive-lending only). "
    "ess/lim/di filled from [to chart] (additive); seeded Key Finding / Design Implication / Causal "
    "Status / Cosmotechnic Implication upgraded to charted text under explicit instruction "
    "(CORE-04 causal->conceptual, CORE-05 causal (lab-bounded), CORE-06 correlational (narrative "
    "review); CI revised per Cosmotechnic-Purist). Verif. set to Charted for 8 rows. Access Type "
    "updated for CORE-04/05/06/08/A1-01 and A1-01 landing URL corrected to verified Urbanomic "
    "publisher page (Brill preview 403) per Evidence Librarian. Locus downgrade tags (supporting "
    "weight) added for CORE-01/02/04/06; A1-01 body-claims verification recorded. No rows removed, "
    "no formulas or sheets touched, no evidence deleted (repair-not-removal convention as in prior "
    "link-replacement patch).",
]
for i, v in enumerate(entry):
    pl.cell(row=pl_append, column=i + 1, value=v)
wb.save(XLSX)
print("PATCH LOG ENTRY ADDED (row", pl_append, ")")

# ---- readback ----
wb2 = load_workbook(XLSX, data_only=True)
ws2 = wb2["Evidence Matrix"]
rows2 = list(ws2.iter_rows(values_only=True))
h2 = [str(c).strip() if c else "" for c in rows2[0]]
c2 = {h: i for i, h in enumerate(h2)}
print("\nREADBACK (Verif. + cell lengths):")
for r in rows2[1:]:
    rid = str(r[c2["ID"]]).strip() if r[c2["ID"]] else ""
    if rid in final:
        kft = str(r[c2["Key Finding / Thesis"]] or "")
        cs = str(r[c2["Causal Status"]] or "")
        print(f"  {rid}: Verif={r[c2['Verif.']]} | KFT={len(kft)}ch | CS={cs[:60]}")
pl2 = wb2["Source Patch Log"]
last = [str(c)[:50] if c else "" for c in list(pl2.iter_rows(values_only=True))[-1]]
print("\nPATCH LOG LAST ROW:", last)
