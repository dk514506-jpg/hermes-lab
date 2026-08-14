#!/usr/bin/env python3
"""Dump batch-2 rows (full columns) to JSON for the role subagents."""
import json
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
TARGET = ["A9-01", "A9-02", "A9-04", "A9-05", "A9-06", "CORE-09", "CORE-10", "CORE-15"]

wb = load_workbook(XLSX, data_only=True)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows(values_only=True))
header = [str(c).strip() if c is not None else "" for c in rows[0]]

out = []
for r in rows[1:]:
    rid = str(r[0]).strip() if r[0] else ""
    if rid in TARGET:
        out.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
with open("/opt/data/Monstare_batch_2_rows.json", "w") as f:
    json.dump(out, f, indent=1, default=str)
print("wrote", len(out), "rows")
