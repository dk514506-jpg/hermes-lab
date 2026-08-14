#!/usr/bin/env python3
"""Extract the 8 spine-batch rows from the Monstare evidence matrix + patch log tail."""
import json, sys

from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
TARGET_IDS = ["CORE-01", "CORE-02", "CORE-04", "CORE-05", "CORE-06", "CORE-08", "A1-01", "HUI-2024"]

wb = load_workbook(XLSX, data_only=True)
print("SHEETS:", wb.sheetnames)

ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows(values_only=True))
header = [str(h).strip() if h is not None else "" for h in rows[0]]
# find ID column
id_col = header.index("ID") if "ID" in header else 0
print("HEADER_COUNT:", len(header))

out = []
for r in rows[1:]:
    if r[id_col] is None:
        continue
    rid = str(r[id_col]).strip()
    if rid in TARGET_IDS:
        out.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})

print("FOUND_ROWS:", len(out), "OF", len(TARGET_IDS))
found = {o.get("ID") for o in out}
print("MISSING:", [i for i in TARGET_IDS if i not in found])

with open("/opt/data/Monstare_batch_1_rows.json", "w") as f:
    json.dump(out, f, indent=1, default=str)

# compact print of the fields relevant to charting
KEYS = ["ID", "Citation", "Readable Source URL", "Source Landing URL", "Access Type",
        "Source Discovery Status", "Source Checked Date", "Verif.", "Area", "WS", "Pri",
        "Function", "Evidence Type", "Domain", "Causal Status", "Cosmo Rel.",
        "Key Finding / Thesis", "Effect Size / Strength", "Limitations",
        "Disconfirming Implication", "H1", "H2", "Design Implication",
        "Cosmotechnic Implication", "Fail Mode", "Artifact Affected", "Open Charge",
        "Staleness Domain", "Re-search Interval", "Last Verified Date", "Re-search By", "Staleness Status"]
for o in out:
    print("=" * 100)
    for k in KEYS:
        v = o.get(k)
        if v is not None and str(v).strip() not in ("", "None"):
            print(f"  {k}: {v}")

# Patch log tail
if "Source Patch Log" in wb.sheetnames:
    pl = wb["Source Patch Log"]
    plrows = list(pl.iter_rows(values_only=True))
    print("\nPATCH LOG (last 12 rows):")
    for r in plrows[-12:]:
        print(" | ".join(str(c) if c is not None else "" for c in r))
