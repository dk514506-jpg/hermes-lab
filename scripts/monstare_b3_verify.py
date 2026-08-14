#!/usr/bin/env python3
"""Independent post-patch verification for batch 3."""
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
TARGET = ["CORE-03", "CORE-11", "CORE-12", "CORE-13", "CORE-14", "CORE-16", "CORE-17", "CORE-18", "CORE-19", "CORE-20"]

wb = load_workbook(XLSX)
print("SHEETS:", wb.sheetnames)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows())
header = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
col = {h: i for i, h in enumerate(header)}

for r in rows[1:]:
    rid = str(r[col["ID"]].value).strip() if r[col["ID"]].value else ""
    if rid not in TARGET:
        continue
    notes = str(r[col["Discovery Notes"]].value or "")
    print(f"--- {rid}")
    print(f"  Verif: {r[col['Verif.']].value}")
    print(f"  Access: {r[col['Access Type']].value}")
    print(f"  EvType: {r[col['Evidence Type']].value}")
    print(f"  Readable: {r[col['Readable Source URL']].value}")
    print(f"  Landing: {str(r[col['Source Landing URL']].value)[:80]}")
    print(f"  Notes tail: ...{notes[-180:]}")
    print(f"  KFT head: {str(r[col['Key Finding / Thesis']].value)[:90]}...")
    print(f"  ES: {str(r[col['Effect Size / Strength']].value)[:60]}")

pl = wb["Source Patch Log"]
print("\nPATCH LOG ROW 7:")
for i, c in enumerate(list(pl.iter_rows())[6], 1):
    v = str(c.value)
    print(f"  col{i}: {v[:220]}{'...' if len(v) > 220 else ''}")
