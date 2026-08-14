#!/usr/bin/env python3
"""Final additive matrix patch: batch-2 charting (workday-protocol spine).
Reads /opt/data/Monstare_batch_2_charting_final.json (post-QC) + access caveats dict."""
import json
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
DATE = "2026-08-13"
final = json.load(open("/opt/data/Monstare_batch_2_charting_final.json"))

COLMAP = {
    "kft": "Key Finding / Thesis", "ess": "Effect Size / Strength", "lim": "Limitations",
    "di": "Disconfirming Implication", "h1": "H1", "h2": "H2",
    "dimpl": "Design Implication", "cimpl": "Cosmotechnic Implication", "cs": "Causal Status",
}
ACCESS_UPDATES = json.load(open("/opt/data/Monstare_batch_2_access_updates.json")) if __import__("os").path.exists("/opt/data/Monstare_batch_2_access_updates.json") else {}
EVTYPE_UPDATES = json.load(open("/opt/data/Monstare_batch_2_evtype_updates.json")) if __import__("os").path.exists("/opt/data/Monstare_batch_2_evtype_updates.json") else {}
CAVEAT_NOTES = {
    "A9-01": "Author-hosted camera-ready PDF (ics.uci.edu); not the publisher version of record; ACM DL paywalled; pages 107-110 per Crossref.",
    "A9-02": "Author-hosted PDF; not the publisher version of record; publisher pagination 113-120 per Crossref; surname Gonzalez per publisher record.",
    "A9-04": "ABSTRACT-LEVEL CHARTING 2026-08-13 (RePEc-verified abstract); full text ScienceDirect paywalled - chart claims only at abstract level; institutional access required for full text.",
    "A9-05": "Full text OA (CC-BY 4.0); PMC article page JS-rendered for curl - read via browser-stack; Frontiers landing open as curl-friendly alternate.",
    "A9-06": "OpenAccess CC-BY 4.0 per RWTH record + Crossref; direct curl PDF fetch blocked by bot challenge - browser-stack required; RWTH 'Review Article' tag is a repository quirk (empirical study); abstract verified via SAGE/IOS landing.",
    "CORE-09": "Course-hosted mirror (weebly) of author-uploaded scan; use published pagination 111-135 (PubMed/Crossref) during charting.",
    "CORE-10": "Course-hosted mirror (ubalt.edu); matches published American Psychologist 57(9):705-717 text.",
    "CORE-15": "ABSTRACT-LEVEL CHARTING 2026-08-13 (PubMed record 9457784); full text institutional (APA PsycNET) - chart only claims supported by the abstract.",
}
DOWNGRADES = {"CORE-09", "A9-02", "A9-05", "A9-06", "CORE-10"}

wb = load_workbook(XLSX)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows())
header = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
col = {h: i for i, h in enumerate(header)}

patched = []
for r in rows[1:]:
    rid = str(r[col["ID"]].value).strip() if r[col["ID"]].value else ""
    if rid not in final:
        continue
    f = final[rid]
    for key, cname in COLMAP.items():
        if key in f and f[key] is not None:
            r[col[cname]].value = f[key]
    r[col["Verif."]].value = f"Charted - {DATE}"
    if rid in ACCESS_UPDATES:
        r[col["Access Type"]].value = ACCESS_UPDATES[rid]
    if rid in EVTYPE_UPDATES:
        r[col["Evidence Type"]].value = EVTYPE_UPDATES[rid]
    notes = str(r[col["Discovery Notes"]].value or "")
    additions = []
    if "CHARTED 2026-08-13 batch 2" not in notes:
        additions.append("CHARTED 2026-08-13 batch 2 (workday-protocol spine; full role QC pass)")
    if rid in CAVEAT_NOTES and CAVEAT_NOTES[rid][:20] not in notes:
        additions.append(CAVEAT_NOTES[rid])
    if rid in DOWNGRADES and "LOCUS" not in notes:
        additions.append("LOCUS 2026-08-13: supporting weight (not load-bearing)")
    if additions:
        r[col["Discovery Notes"]].value = (notes + "; " if notes else "") + "; ".join(additions)
    r[col["Source Checked Date"]].value = DATE
    patched.append(rid)

wb.save(XLSX)
print("PATCHED:", patched)

pl = wb["Source Patch Log"]
pl_append = pl.max_row + 1
entry = [
    DATE,
    "Additive patch - batch 2 charting (QC-complete)",
    ", ".join(patched),
    "None (no schema change; existing charting columns filled)",
    "0",
    "0",
    "Second charting pass (8-row workday-protocol spine) with role QC (Evidence Librarian, "
    "Methodologist, Phenomenologist, Cosmotechnic-Purist, Ethics & Cosmotechnic Auditor, "
    "Data & Instrumentation Steward, Locus). Citation-is-not-evidence respected. A9-04, A9-06, "
    "CORE-15 charted at abstract level (full text institutional/blocked - caveats recorded); "
    "A9-05 via open-access full text; A9-01/A9-02/CORE-09/CORE-10 from full-text PDFs. "
    "Verif. set to Charted for 8 rows; Access Type caveats updated where the Librarian directed. "
    "No rows removed, no formulas or sheets touched, additive only.",
]
for i, v in enumerate(entry):
    pl.cell(row=pl_append, column=i + 1, value=v)
wb.save(XLSX)
print("PATCH LOG ENTRY ADDED (row", pl_append, ")")

wb2 = load_workbook(XLSX, data_only=True)
ws2 = wb2["Evidence Matrix"]
rows2 = list(ws2.iter_rows(values_only=True))
h2 = [str(c).strip() if c else "" for c in rows2[0]]
c2 = {h: i for i, h in enumerate(h2)}
print("\nREADBACK:")
for r in rows2[1:]:
    rid = str(r[c2["ID"]]).strip() if r[c2["ID"]] else ""
    if rid in final:
        print(f"  {rid}: Verif={r[c2['Verif.']]} | KFT={len(str(r[c2['Key Finding / Thesis']] or ''))}ch | CS={str(r[c2['Causal Status']])[:55]}")
