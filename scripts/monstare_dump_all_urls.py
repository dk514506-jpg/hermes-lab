#!/usr/bin/env python3
"""Dump ALL evidence-matrix rows' URL fields for the link audit."""
import json
from openpyxl import load_workbook

XLSX = "/opt/data/Monstare_Evidence_Matrix_Source_Links_v3_Staleness_Patched_artifact.xlsx"
wb = load_workbook(XLSX, data_only=True)
ws = wb["Evidence Matrix"]
rows = list(ws.iter_rows(values_only=True))
header = [str(h).strip() if h is not None else "" for h in rows[0]]

KEYS = ["ID", "Citation", "Readable Source URL", "Source Landing URL", "Access Type",
        "Source Discovery Status", "Discovery Notes", "Suggested Search Query",
        "Scholar/Web Search URL", "Verif.", "Area", "Pri", "Staleness Domain",
        "Re-search Interval", "Staleness Status"]

out = []
for r in rows[1:]:
    d = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
    if d.get("ID") is None:
        continue
    out.append({k: d.get(k) for k in KEYS})

with open("/opt/data/Monstare_all_rows_urls.json", "w") as f:
    json.dump(out, f, indent=1, default=str)

n = len(out)
no_readable = [o["ID"] for o in out if not (o.get("Readable Source URL") or "").strip()]
no_landing = [o["ID"] for o in out if not (o.get("Source Landing URL") or "").strip()]
print(f"ROWS: {n}")
print(f"ROWS WITHOUT READABLE URL: {len(no_readable)} {no_readable[:20]}")
print(f"ROWS WITHOUT LANDING URL: {len(no_landing)} {no_landing[:20]}")

readable = [o for o in out if (o.get("Readable Source URL") or "").strip()]
landing = [o for o in out if (o.get("Source Landing URL") or "").strip()]
print(f"ROWS WITH READABLE URL: {len(readable)}")
print(f"ROWS WITH LANDING URL: {len(landing)}")

all_urls = set()
for o in out:
    for k in ("Readable Source URL", "Source Landing URL"):
        u = (o.get(k) or "").strip()
        if u:
            all_urls.add(u)
print(f"UNIQUE URLS (both columns): {len(all_urls)}")

# sample of access types
from collections import Counter
print("ACCESS TYPE COUNTS:", Counter((o.get("Access Type") or "").strip() or "(blank)" for o in out).most_common(20))
print("DISCOVERY STATUS COUNTS:", Counter((o.get("Source Discovery Status") or "").strip() or "(blank)" for o in out).most_common(20))

# quick host breakdown of readable URLs
from urllib.parse import urlparse
hosts = Counter(urlparse(u).netloc for u in all_urls)
print("HOSTS (top 25):")
for h, c in hosts.most_common(25):
    print(f"  {c:4d}  {h}")
